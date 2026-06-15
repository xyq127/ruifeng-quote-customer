#!/usr/bin/env python3
"""
张紧轮/惰轮报价清单全量清洗脚本（含知识库）
适用于有盖茨GTA码的发电机张紧器/惰轮/过渡轮报价清单。
修改 SOURCE_PATH 和 OUTPUT_PATH 后运行。

依赖：openpyxl
"""
import csv, sys

# ====== 修改这里 ======
SOURCE_PATH = "/mnt/c/Users/Lenovo/Documents/雷迪克/张紧轮/涨紧轮订单.xlsx"
OUTPUT_PATH = "/mnt/c/Users/Lenovo/Documents/雷迪克/张紧轮/涨紧轮报价底表_清洗结果.csv"
# =====================

# ====== 盖茨 GTA → OE号/发动机映射（核心105条） ======
GATES_DB = {
    # 现代/起亚
    "GTA1043": {"oe": "25281-2B000, 25281-2B010", "engine": "G4FA/G4FC Gamma 1.4/1.6L"},
    "GTA1245": {"oe": "25281-2B000", "engine": "G4FA/G4FC Gamma 1.4/1.6L"},
    "GTA5033A": {"oe": "25281-2G000", "engine": "G4KD Theta-II 2.0L"},
    "GTA5018": {"oe": "25281-2B010", "engine": "G4FA/G4FC Gamma"},
    "GTA5022": {"oe": "97617-22000, 97617-23030", "engine": "G4EA/G4ED/G4GC"},
    
    # 大众集团
    "GTA1004": {"oe": "06B903315A, 06B903315C", "engine": "EA113 1.8L/1.8T"},
    "GTA1073": {"oe": "04E903315B, 04E903315C", "engine": "EA211 1.6L MPI"},
    "GTA1287": {"oe": "04E903315D, 04E903315E", "engine": "EA211 1.4T/1.5L"},
    "GTA1040": {"oe": "03C903315A, 03C903315B", "engine": "EA111 1.4T TSI"},
    "GTA1134A": {"oe": "06J903315C, 06J903315D", "engine": "EA888 Gen2 1.8T/2.0T"},
    "GTA1064": {"oe": "06E903315A, 06E903315E", "engine": "CRE/CJT 3.0T V6"},
    "GTA1120": {"oe": "06E903315A, 06E903315E", "engine": "3.0T V6 TFSI"},
    "GTA1139": {"oe": "06H903315, 06K903315A", "engine": "EA888 Gen3 2.0L"},
    "GTA5004": {"oe": "03C903315C", "engine": "EA111 1.4T 惰轮"},
    "GTA5175": {"oe": "03C903315C", "engine": "EA111 1.6L 过渡轮"},
    "GTA5085": {"oe": "06E903315A", "engine": "BDW 2.4L V6 惰轮"},
    "GTA5086": {"oe": "06H903315, 06K903315A", "engine": "EA888 1.8T/2.0T 惰轮"},
    
    # 丰田/雷克萨斯
    "GTA1078A": {"oe": "16620-0P010, 16620-31070", "engine": "1AR-FE 2.7L L4"},
    "GTA1087": {"oe": "16620-0H010, 16620-28041", "engine": "2AZ-FE 2.4L L4"},
    "GTA1042A": {"oe": "16620-46030, 16620-46040", "engine": "3GR-FE 3.0L V6"},
    "GTA1118A": {"oe": "16620-28050, 16620-28041", "engine": "1AZ-FE 2.0L L4"},
    "GTA1362": {"oe": "16620-0Q010, 16620-0Q020", "engine": "2SZ-FE/1NR-FE 1.3L"},
    "GTA1187": {"oe": "16620-50011, 16620-50020", "engine": "2UZ-FE 4.7L V8"},
    "GTA1395": {"oe": "16620-31070, 16620-F4020", "engine": "6AR-FSE/M20A-FKS 2.0L"},
    "GTA5100": {"oe": "16620-31070, 16620-38050", "engine": "1GR-FE/2AR-FE"},
    "GTA5088": {"oe": "16620-50011, 16620-50020", "engine": "1GR-FE 4.0L V6"},
    "GTA5145A": {"oe": "16620-28050, 16620-28041", "engine": "1AZ-FE/2AZ-FE 惰轮"},
    "GTA1314": {"oe": "16620-31070, 16620-F4020", "engine": "8AR-FTS 2.0T"},
    
    # 本田
    "GTA1088": {"oe": "31110-RAA-A01, 31110-RNA-A01", "engine": "R20A3/R18A1"},
    "GTA1317": {"oe": "31110-R1A-A01, 31110-R1A-004", "engine": "R18Z 1.8L"},
    "GTA1133": {"oe": "31110-R1A-A01, 31110-5T0-A01", "engine": "L15B 1.5L地球梦"},
    "GTA5158": {"oe": "31110-PNA-006, 31110-PNA-004", "engine": "L13A 1.4L 空调轮"},
    
    # 日产
    "GTA1199A": {"oe": "11955-JA00A, 11955-JD000", "engine": "MR20DE 2.0L"},
    "GTA1069": {"oe": "11955-JA00A, 11955-JD000", "engine": "MR20DE 2.0L"},
    "GTA1070": {"oe": "11955-ED000, 11955-ED00A", "engine": "HR16DE 1.6L"},
    "GTA1050": {"oe": "11955-JA00A, 11955-5RB0A", "engine": "MR18DE/MR20DE"},
    "GTA1290": {"oe": "11955-ED000, 11955-ED00A", "engine": "HR16DE 1.6L"},
    "GTA1373": {"oe": "11955-5RB0A, 11955-5RB1A", "engine": "HR16DE Gen3"},
    "GTA1393": {"oe": "11955-5RB0A, 11955-5RB1A", "engine": "MR20DD 2.0L"},
    "GTA5226": {"oe": "11955-ED000, 11955-ED00A", "engine": "HR16DE 惰轮"},
    
    # 福特
    "GTA1011": {"oe": "3M5Q-6K255-AA, 3M5Q-6K255-AB", "engine": "Duratec-HE"},
    "GTA5081": {"oe": "3M5Q-6K255-AA, CV6Q-6K255-AA", "engine": "EcoBoost 1.6L"},
    "GTA1461": {"oe": "BB5Z-6K255-A, DB5Z-6K255-A", "engine": "3.5L V6"},
    
    # 标致/雪铁龙
    "GTA1033": {"oe": "5751.G3, 5751.R0", "engine": "EW10A/EW12"},
    "GTA5014": {"oe": "5751.G3", "engine": "EW10A 惰轮"},
    "GTA5352": {"oe": "5751.43", "engine": "EW10J4 单轮"},
    
    # 马自达
    "GTA1143": {"oe": "PE01-15-980, PE03-15-980", "engine": "PE Skyactiv-G 2.0"},
    "GTA1024": {"oe": "F2R1-15-980, F231-15-980", "engine": "Z6/MZ-CD"},
    "GTA5008": {"oe": "F2R1-15-980, F231-15-980", "engine": "Z6/MZ-CD 惰轮"},
    
    # 奔驰
    "GTA1074": {"oe": "2712001070, 2732000170", "engine": "M273 5.5L V8"},
    "GTA1101": {"oe": "2722001070, 2762000070", "engine": "M272/M276"},
    "GTA5045A": {"oe": "2732000170", "engine": "M273 导向轮"},
    "GTA5242": {"oe": "2712001070", "engine": "M271 1.8L"},
    "GTA5243": {"oe": "2712001070", "engine": "M271 Kompressor"},
    
    # 宝马
    "GTA1051": {"oe": "11287507964, 11287571637", "engine": "N46/N52/N20"},
    "GTA1334": {"oe": "11287507964, 11287571637", "engine": "N46/N20"},
    "GTA5234": {"oe": "11287507964", "engine": "N52/N54 惰轮"},
    "GTA5239": {"oe": "11287507964", "engine": "N46 惰轮"},
    
    # 长城
    "GTA1148A": {"oe": "1026200-E01", "engine": "GW4D20 2.0T柴油"},
    "GTA1298": {"oe": "1026200-E01, 1002100-E01", "engine": "GW4C20 2.0T"},
    "GTA1301": {"oe": "1026200-E01", "engine": "4G63S4T 2.0T"},
    "GTA1342": {"oe": "1026200-E01", "engine": "GW4C20 2.0T"},
    "GTA1371": {"oe": "1026200-E01", "engine": "GW4B15 1.5T"},
    "GTA5094A": {"oe": "1026200-E01", "engine": "GW4D20 惰轮"},
    "GTA5095A": {"oe": "1026200-E01", "engine": "GW4D20 纵置惰轮"},
    
    # 奇瑞
    "GTA1008": {"oe": "S11-3701100, A11-3701100", "engine": "ACTECO SQR481F"},
    "GTA1053": {"oe": "SQRD4T15, SQRD4T15B", "engine": "SQRE4T15 1.5T"},
    "GTA5282": {"oe": "S11-3701100", "engine": "SQR481F 惰轮"},
    "GTA5003": {"oe": "A11-3701100", "engine": "ACTECO 过渡轮"},
    "GTA5036": {"oe": "", "engine": "SQRE4G15 1.5L 惰轮"},
    
    # 吉利
    "GTA1255": {"oe": "1016000326, 1016000327", "engine": "JL4G15 1.5L"},
    "GTA1276": {"oe": "1016000336, 1016000506", "engine": "MR479QA/JL4G15"},
    "GTA5189": {"oe": "", "engine": "MR479QA 惰轮"},
    
    # 通用
    "GTA1060": {"oe": "12560678, 12620831", "engine": "LZC/LFW 3.0L V6"},
    "GTA1076": {"oe": "12560678, 12620831", "engine": "LZC/LY7 3.0/3.6L"},
    "GTA5039": {"oe": "12560678, 12620831", "engine": "LZC/LFW 惰轮"},
    
    # 克莱斯勒
    "GTA1163": {"oe": "4891990AA, 4891990AB", "engine": "World Engine"},
    "GTA5076": {"oe": "04891764AA, 4891990AA", "engine": "World Engine 过渡轮"},
    "GTA5299": {"oe": "04891764AA", "engine": "World Engine 小过渡轮"},
    
    # 其他
    "GTA1241": {"oe": "24547902, 23888680", "engine": "LMH/L2B 五菱"},
    "GTA1044": {"oe": "23888680, 24438093", "engine": "L2B P-TEC 宝骏730"},
    "GTA1160": {"oe": "12651236, 24483200", "engine": "LMH 1.2L 宝骏310"},
    "GTA5037": {"oe": "", "engine": "SQRE4T15C 1.5T 惰轮"},
    "GTA5084": {"oe": "24547902", "engine": "LL5 1.5T 惰轮"},
    "GTA5319": {"oe": "", "engine": "JL473ZQ 1.5T 惰轮"},
    "GTA1349": {"oe": "", "engine": "F15C"},
    "GTA1144": {"oe": "S180000400", "engine": "JL486ZQ2 1.8T"},
    "GTA1426": {"oe": "", "engine": "2.8T/2.0T/1.9T 大通"},
    "GTA1369": {"oe": "", "engine": "1.9T柴油 大通G10"},
    "GTA1433": {"oe": "", "engine": "JL473ZQ 1.5T/JL473ZQ5 1.4T"},
    "GTA1492": {"oe": "LR025377, LR079235", "engine": "Ingenium AJ20 2.0T"},
    "GTA1330": {"oe": "31316704, 30757185", "engine": "B4204T/B4154T 沃尔沃"},
    "GTA1427": {"oe": "", "engine": "JE4D25 2.5T柴油"},
    "GTA1387": {"oe": "", "engine": "Duratorq 2.2T"},
    "GTA1142A": {"oe": "6C11-6K255-AA", "engine": "Duratorq 全顺V348"},
    "GTA5280": {"oe": "6C11-6K255-AA", "engine": "Duratorq 惰轮"},
    "GTA5215": {"oe": "MN982280, MR128261", "engine": "6G72/6G75 帕杰罗"},
    "GTA5262": {"oe": "MN982280", "engine": "4B11 惰轮"},
    "GTA5106": {"oe": "", "engine": "1.8T/1.9T"},
    "GTA5182": {"oe": "", "engine": "1.9T柴油 大通惰轮"},
    "GTA5314": {"oe": "", "engine": "JE4D25/RZ4E 惰轮"},
    "GTA5316": {"oe": "", "engine": "JE4D25 惰轮"},
    "GTA5317": {"oe": "", "engine": "JE4D25 柴油惰轮"},
    "GTA5091": {"oe": "", "engine": "JL486ZQ2 惰轮"},
    "GTA5092": {"oe": "", "engine": "JL486ZQ2 惰轮2"},
}

# ====== 行话→发动机/OE映射 ======
VEHICLE_DB = {
    "本田雅阁2.4L": {"oe": "31110-RAA-A01, 31110-RBA-004", "engine": "K24Z (八代2008-2012)", "note": "按八代雅阁2.4L K24Z匹配，不适用九代2013+款"},
    "朗逸1.6": {"oe": "03C903315C", "engine": "EA111 (1.6L MPI)", "note": "按EA111匹配，不适用EA211新款"},
    "波罗劲情劲取": {"oe": "03C903315C", "engine": "EA111 (1.4/1.6L)", "note": "按Polo劲情劲取EA111匹配"},
    "标致207 307 1.6": {"oe": "5751.43, 5751.42", "engine": "TU5JP4 (1.6L)", "note": "按TU5JP4匹配，不适用2.0L"},
    "索纳塔8": {"oe": "25281-2G000", "engine": "G4KD Theta-II/NU 2.0L", "note": "按索八Theta-II 2.0L匹配"},
    "凯美瑞2.4": {"oe": "16620-0H010, 16620-28041", "engine": "2AZ-FE (2.4L)", "note": "按2.4L 2AZ-FE匹配，不适用2.0L"},
    "汉兰达2.7": {"oe": "16620-0P010, 16620-31070", "engine": "1AR-FE (2.7L)", "note": "按汉兰达2.7L匹配"},
    "霸道2700": {"oe": "16620-31070, 16620-69035", "engine": "2TR-FE (2.7L)", "note": "按霸道2700 2.7L匹配"},
    "皇冠3.0": {"oe": "16620-46030", "engine": "3GR-FE (3.0L V6)", "note": "按皇冠3.0L匹配"},
    "A6L2.4": {"oe": "06E903315A", "engine": "BDW (2.4L V6)", "note": "按A6L C6 2.4L BDW匹配"},
    "捷达王": {"oe": "06B903315A, 06B903315C", "engine": "EA113 (1.8L/1.8T)", "note": "按EA113匹配"},
    "新桑塔纳": {"oe": "04E903315B, 04E903315C", "engine": "EA211 (1.6L)", "note": "按EA211 1.6L匹配"},
    "天籁2.5": {"oe": "11955-JD000, 11955-JA00A", "engine": "VQ25DE", "note": "按天籁2.5 V6匹配"},
    "天籁2.0": {"oe": "11955-JA00A", "engine": "MR20DE", "note": "按天籁2.0匹配"},
    "逍客": {"oe": "11955-JA00A", "engine": "MR20DE (2.0L)", "note": "按逍客/奇骏2.0匹配"},
    "轩逸1.6": {"oe": "11955-ED000, 11955-5RB0A", "engine": "HR16DE", "note": "按轩逸1.6 HR16匹配"},
    "蒙迪欧": {"oe": "3M5Q-6K255-AA", "engine": "Duratec-HE/EcoBoost", "note": "按蒙迪欧系列匹配"},
    "福克斯": {"oe": "3M5Q-6K255-AA", "engine": "Duratec-HE (1.8/2.0L)", "note": "按福克斯匹配"},
    "翼虎": {"oe": "CV6Q-6K255-AA", "engine": "EcoBoost", "note": "按翼虎系列匹配"},
    "君越2.4": {"oe": "12560678, 12620831", "engine": "LE5/LEA (2.4L)", "note": "按君越2.4 EcoTec匹配"},
    "GL8": {"oe": "12560678, 12620831", "engine": "LZC/LFW (3.0L)", "note": "按GL8系列匹配"},
}

def load_data(path):
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for i in range(2, ws.max_row + 1):
        rows.append({
            'prod_code': str(ws.cell(i, 2).value or '').strip(),
            'prod_name': str(ws.cell(i, 3).value or '').strip(),
            'vehicle': str(ws.cell(i, 4).value or '').strip(),
            'qty': ws.cell(i, 5).value,
            'gates': str(ws.cell(i, 6).value or '').strip(),
        })
    return rows

def main():
    rows = load_data(SOURCE_PATH)
    output = []
    
    for r in rows:
        gates = r['gates']
        vehicle = r['vehicle']
        
        std_oe = ""
        engine_model = ""
        match_note = ""
        confidence = "Low"
        
        # Condition A
        if gates and gates not in ('None', ''):
            if gates in GATES_DB:
                info = GATES_DB[gates]
                std_oe = info['oe']
                engine_model = info['engine']
                match_note = f"盖茨{gates}精准匹配"
                confidence = "High" if std_oe else "Medium"
            else:
                match_note = f"盖茨码{gates}未知，需核对"
                confidence = "Low"
        
        # Condition B
        if confidence != "High":
            for vkey, vinfo in VEHICLE_DB.items():
                if vkey[:6] in vehicle or vehicle[:6] in vkey:
                    std_oe = vinfo['oe'] if not std_oe else std_oe
                    engine_model = vinfo['engine']
                    match_note = f"行话解析：{vinfo['note']}"
                    confidence = "Medium" if std_oe else "Low"
                    break
            else:
                # Brand-level fallback
                if '长安' in vehicle or '比亚迪' in vehicle or '江淮' in vehicle:
                    match_note = "需客户提供原厂OE或车架号，自主品牌不易交叉验证"
                elif '惰轮' in r['prod_name']:
                    match_note = "惰轮类，轴承尺寸匹配即可，需客户确认OE"
                else:
                    match_note = "需客户提供原厂OE或车架号"
        
        output.append({
            '行号': '', '商品编号': r['prod_code'], '商品名称': r['prod_name'],
            '适用车型描述': r['vehicle'], '数量': r['qty'], '盖茨码': gates,
            '标准OE号': std_oe, '发动机型号': engine_model,
            '匹配备注(给客户)': match_note, '置信度': confidence
        })
    
    with open(OUTPUT_PATH, 'w', newline='', encoding='utf-8-sig') as f:
        w = csv.DictWriter(f, fieldnames=[
            '行号','商品编号','商品名称','适用车型描述','数量','盖茨码',
            '标准OE号','发动机型号','匹配备注(给客户)','置信度'
        ])
        w.writeheader()
        w.writerows(output)
    
    h = sum(1 for r in output if r['置信度'] == 'High')
    m = sum(1 for r in output if r['置信度'] == 'Medium')
    l = sum(1 for r in output if r['置信度'] == 'Low')
    print(f"完成！{len(output)}行：High={h} Medium={m} Low={l}")
    print(f"输出：{OUTPUT_PATH}")

if __name__ == '__main__':
    main()
